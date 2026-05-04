#!/usr/bin/env python3
"""Run an AQL query and export the JSON result to a flat Excel sheet.

The script is generic: provide any AQL query and it converts nested JSON rows
into a columnar format by flattening nested objects with underscore-separated
column names. Lists are preserved as JSON strings.

PyCharm usage (no args)
- Uses `EXAMPLE_QUERY` against the default settings file.
- Writes a timestamped Excel file in `Analysis/`.

Examples
- Use built-in query:
  python Analysis/aql_to_excel.py --use-example --limit 100
- Use inline query:
  python Analysis/aql_to_excel.py --query "FOR a IN assets LIMIT 10 RETURN {uuid: a._key, naam: a.AIMNaamObject_naam}"
- Use query file:
  python Analysis/aql_to_excel.py --query-file Analysis/lsb_keuring_overview.aql --bind-vars-json '{"limit": 200}'
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from API.APIEnums import Environment
from Analysis.export_keuringsinfo import _create_db_from_settings, _load_settings

DEFAULT_SETTINGS_PATH = Path("/home/davidlinux/Documenten/AWV/resources/settings_SyncToArangoDB.json")
DEFAULT_SHEET_NAME = "result"

EXAMPLE_QUERY = (
    "FOR a IN assets\n"
    "  LIMIT @limit\n"
    "  RETURN {\n"
    "    uuid: a._key,\n"
    "    type: a.`@type`,\n"
    "    naam: a.AIMNaamObject_naam,\n"
    "    naampad: a.NaampadObject_naampad,\n"
    "    isActief: a.AIMDBStatus_isActief\n"
    "  }"
)


def _normalize_key_part(key: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-z_]+", "_", key).strip("_")
    return normalized or "value"


def _flatten_value(value: Any, *, prefix: str) -> dict[str, Any]:
    """Flatten nested dicts; preserve lists as JSON strings."""
    flat: dict[str, Any] = {}

    def visit(current: Any, path: str) -> None:
        if isinstance(current, dict):
            if not current:
                flat[path] = "{}"
                return
            for child_key, child_value in current.items():
                next_path = f"{path}_{_normalize_key_part(str(child_key))}"
                visit(child_value, next_path)
            return

        if isinstance(current, list):
            flat[path] = json.dumps(current, ensure_ascii=False)
            return

        flat[path] = current

    visit(value, prefix)
    return flat


def _flatten_row(row: Any) -> dict[str, Any]:
    if isinstance(row, dict):
        flattened: dict[str, Any] = {}
        for key, value in row.items():
            col = _normalize_key_part(str(key))
            if isinstance(value, dict):
                flattened.update(_flatten_value(value, prefix=col))
            elif isinstance(value, list):
                flattened[col] = json.dumps(value, ensure_ascii=False)
            else:
                flattened[col] = value
        return flattened

    return {"value": row}


def _sanitize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _autofit_sheet_columns(ws: Any, *, min_width: int = 10, max_width: int = 80, padding: int = 2) -> None:
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        first = next(iter(col_cells), None)
        if first is None or getattr(first, "column", None) is None:
            continue

        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))

        width = min(max(min_width, max_len + padding), max_width)
        ws.column_dimensions[get_column_letter(first.column)].width = width


def _parse_bind_vars(bind_vars_json: str) -> dict[str, Any]:
    if not bind_vars_json.strip():
        return {}
    parsed = json.loads(bind_vars_json)
    if not isinstance(parsed, dict):
        raise ValueError("--bind-vars-json must decode to a JSON object")
    return parsed


def _resolve_query(args: argparse.Namespace) -> str:
    if args.query:
        return args.query
    if args.query_file:
        return args.query_file.read_text(encoding="utf-8")
    # PyCharm-friendly default: run the built-in example when no query is provided.
    return EXAMPLE_QUERY


def fetch_rows(
    db: Any,
    *,
    query: str,
    bind_vars: dict[str, Any],
    max_runtime_seconds: int,
    batch_size: int,
    ttl: int,
    limit: int | None,
) -> list[Any]:
    cursor = db.aql.execute(
        query,
        bind_vars=bind_vars,
        batch_size=batch_size,
        ttl=ttl,
        max_runtime=max_runtime_seconds,
        stream=True,
    )

    rows: list[Any] = []
    for idx, row in enumerate(cursor):
        if limit is not None and idx >= limit:
            break
        rows.append(row)
    return rows


def export_rows_to_excel(rows: list[Any], out_path: Path, *, sheet_name: str) -> tuple[int, int]:
    from openpyxl import Workbook

    flattened_rows = [_flatten_row(row) for row in rows]

    headers_set: set[str] = set()
    for row in flattened_rows:
        headers_set.update(row.keys())
    headers = sorted(headers_set)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers:
        ws.append(headers)
        for row in flattened_rows:
            ws.append([_sanitize_cell(row.get(header)) for header in headers])
    else:
        ws.append(["info"])
        ws.append(["Query returned no rows"])

    _autofit_sheet_columns(ws)
    wb.save(out_path)
    return len(flattened_rows), len(headers)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run AQL and export results to Excel")
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS_PATH, help="Path to settings json")
    parser.add_argument("--env", type=str, default="PRD", choices=[e.name for e in Environment])

    parser.add_argument("--query", type=str, default=None, help="Inline AQL query")
    parser.add_argument("--query-file", type=Path, default=None, help="Path to .aql file")
    parser.add_argument("--use-example", action="store_true", help="Force use of built-in EXAMPLE_QUERY")
    parser.add_argument("--bind-vars-json", type=str, default="{}", help="JSON object of AQL bind vars")

    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).with_name(f"aql_export_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
        help="Output .xlsx path",
    )
    parser.add_argument("--sheet-name", type=str, default=DEFAULT_SHEET_NAME)
    parser.add_argument("--limit", type=int, default=None, help="Optional cap on fetched rows")
    parser.add_argument("--max-runtime-seconds", type=int, default=300)
    parser.add_argument("--batch-size", type=int, default=2000)
    parser.add_argument("--ttl", type=int, default=600)

    args = parser.parse_args(argv)

    if args.use_example:
        args.query = EXAMPLE_QUERY
        args.query_file = None
    elif args.query and args.query_file:
        raise ValueError("Use either --query or --query-file, not both")

    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if not args.settings.exists():
        raise FileNotFoundError(f"Settings file not found: {args.settings}")

    query = _resolve_query(args)
    bind_vars = _parse_bind_vars(args.bind_vars_json)

    # Example query expects @limit bind var when not provided by user.
    if query == EXAMPLE_QUERY and "limit" not in bind_vars:
        bind_vars["limit"] = 100

    settings = _load_settings(args.settings)
    db = _create_db_from_settings(settings, env=Environment[args.env])

    rows = fetch_rows(
        db,
        query=query,
        bind_vars=bind_vars,
        max_runtime_seconds=args.max_runtime_seconds,
        batch_size=args.batch_size,
        ttl=args.ttl,
        limit=args.limit,
    )

    row_count, column_count = export_rows_to_excel(rows, args.out, sheet_name=args.sheet_name)
    print(f"AQL returned {row_count} rows")
    print(f"Excel columns: {column_count}")
    print(f"Wrote Excel file to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

