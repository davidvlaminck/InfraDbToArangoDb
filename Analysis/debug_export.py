#!/usr/bin/env python3
"""
Debug/export helper that bundles smaller Analysis scripts.

Provides a small CLI to:
- generate the keuringsinfo Excel (reuses `export_keuringsinfo.export_to_excel`)
- compare Python-calculated pivot totals with the Pivot sheet in the generated Excel
- run a pivot-debug that prints ignored-count reasons (from debug_pivot_counts)
- a smoke mode to run a small end-to-end check (limited rows)

This file purposely re-uses the canonical logic in `Analysis/export_keuringsinfo.py` so
there's only one source of truth for parsing, pivoting and writing Excel.

Usage examples:
  python Analysis/debug_export.py --mode generate --settings settings.json --out ./out.xlsx
  python Analysis/debug_export.py --mode compare --settings settings.json --out ./out.xlsx
  python Analysis/debug_export.py --mode both --settings settings.json
  python Analysis/debug_export.py --mode smoke --settings settings.json --limit 50

"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
import tempfile
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, Tuple

# Ensure repository root is importable so `from Analysis import export_keuringsinfo` works
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from API.APIEnums import Environment
from Analysis.export_keuringsinfo import (
    PIVOT_SHEET,
    PIVOT_ALL_SHEET,
    _build_pivot,
    _create_db_from_settings,
    _load_settings,
    _parse_iso_date,
    export_to_excel,
    fetch_records,
)


def _default_out_path() -> Path:
    return Path(__file__).with_name(f"keuringsinfo_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


def generate_excel(
    settings_path: Path,
    env: Environment,
    ls_short_uri: str,
    lsdeel_short_uri: str,
    out_path: Path | None = None,
    limit: int | None = None,
) -> Tuple[Path, list]:
    """Fetch records and write Excel. Returns (out_path, records_list).

    Note: may raise on DB or IO errors; caller should handle.
    """
    if out_path is None:
        out_path = _default_out_path()

    settings = _load_settings(settings_path)
    db = _create_db_from_settings(settings, env=env)

    records = fetch_records(
        db,
        ls_short_uri=ls_short_uri,
        lsdeel_short_uri=lsdeel_short_uri,
        limit=limit,
    )

    export_to_excel(records, out_path)
    return out_path, records


def _read_pivot_totals_from_excel(path: Path, sheet: str = PIVOT_SHEET) -> Tuple[list[str], dict[str, int]]:
    """Read the Pivot sheet totals row from the Excel file.

    Returns (cols_order, totals_map) where cols_order is the result-columns order (matching export)
    and totals_map maps column-name -> int(total).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Pivot sheet '{sheet}' not found in workbook: {path}")

    ws = wb[sheet]

    # header is expected as: toezichtgroep, <cols...>, Totaal
    header = None
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if row and row[0] and str(row[0]).strip().lower() == 'toezichtgroep':
            header = list(row)
            break
    if header is None:
        # fallback to first row
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    # result columns are header[1:-1]
    cols = [c for c in header[1:-1]]

    # find the 'Totaal' row
    totals_row = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = row[0]
        if first and str(first).strip().lower() == 'totaal':
            totals_row = list(row)
            break

    if totals_row is None:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError(f"No rows found in pivot sheet: {sheet}")
        totals_row = list(all_rows[-1])

    # Map cols to totals: cols correspond to positions 1..n
    totals_map: dict[str, int] = {}
    for idx, col in enumerate(cols, start=1):
        try:
            v = totals_row[idx]
        except Exception:
            v = 0
        totals_map[str(col).strip()] = int(v or 0)

    return cols, totals_map


def build_python_pivot_totals(records: Iterable, *, cutoff: dt.date, include_not_meegenomen: bool = False) -> Tuple[list[str], dict[str, int]]:
    """Use the canonical _build_pivot to compute totals aggregated across groups.

    Returns (cols_order, totals_map) where totals_map[col] = total across all groups.
    """
    cols, counters = _build_pivot(records, cutoff=cutoff, include_not_meegenomen=include_not_meegenomen)
    grand: dict[str, int] = {c: 0 for c in cols}
    for grp_counter in counters.values():
        for c in cols:
            grand[c] += int(grp_counter.get(c, 0))
    return cols, grand


def compare_records_vs_excel(records: Iterable, excel_path: Path, *, include_not_meegenomen: bool = False) -> Dict[str, Dict[str, Any]]:
    """Compare python-calculated pivot totals with Excel pivot totals.

    Returns a dict mapping column -> {py:int, excel:int, diff:int}
    """
    cutoff = dt.date(2021, 1, 1)
    cols_py, py_totals = build_python_pivot_totals(records, cutoff=cutoff, include_not_meegenomen=include_not_meegenomen)
    sheet = PIVOT_ALL_SHEET if include_not_meegenomen else PIVOT_SHEET
    cols_xl, xl_totals = _read_pivot_totals_from_excel(excel_path, sheet=sheet)

    # normalize col names (strip) and ensure order
    cols = cols_py

    report: dict[str, dict[str, Any]] = {}
    for c in cols:
        py = int(py_totals.get(c, 0))
        xl = int(xl_totals.get(c, 0))
        report[c] = {"py": py, "excel": xl, "diff": py - xl}
    return report


def run_pivot_debug(settings_path: Path, env: Environment, ls_short_uri: str, lsdeel_short_uri: str, limit: int | None = None) -> int:
    """Re-implements the debug_pivot_counts behaviour and prints counts and reasons."""
    settings = _load_settings(settings_path)
    db = _create_db_from_settings(settings, env=env)

    recs = fetch_records(db, ls_short_uri=ls_short_uri, lsdeel_short_uri=lsdeel_short_uri, limit=limit)

    cutoff = dt.date(2021, 1, 1)

    stats = Counter()
    ignored = Counter()
    res_counter = Counter()

    for r in recs:
        stats["records"] += 1
        stats[f"type:{r.type}"] += 1
        stats[f"match:{r.match}"] += 1

        has_ins = bool(r.datum_laatste_keuring or r.resultaat_keuring)
        stats["has_ins"] += int(has_ins)

        d = _parse_iso_date(r.datum_laatste_keuring)
        if d is None:
            ignored["no_date"] += 1
            continue
        if d <= cutoff:
            ignored["date_le_cutoff"] += 1
            continue

        rr = (r.resultaat_keuring or "").strip()
        if not rr:
            ignored["blank_result"] += 1
            continue

        res_counter[rr] += 1
        stats["counted"] += 1

    print("=== BASIC ===")
    for k, v in stats.most_common():
        print(f"{k}: {v}")

    print("\n=== IGNORED (why not counted in pivot) ===")
    for k, v in ignored.most_common():
        print(f"{k}: {v}")

    print("\n=== COUNTED RESULTS (top 20) ===")
    for k, v in res_counter.most_common(20):
        print(f"{k}: {v}")

    return 0


def run_quick_stats(settings_path: Path, env: Environment, ls_short_uri: str, lsdeel_short_uri: str, limit: int | None = None, sample: int = 10) -> int:
    """Print a quick console sample and counters similar to main_export_keuringsinfo.

    Useful for a quick sanity check without writing Excel.
    """
    settings = _load_settings(settings_path)
    db = _create_db_from_settings(settings, env=env)

    records = fetch_records(db, ls_short_uri=ls_short_uri, lsdeel_short_uri=lsdeel_short_uri, limit=limit)

    print(f"Fetched {len(records)} records")
    # type and match counts
    t_counts = Counter(r.type for r in records)
    m_counts = Counter(r.match for r in records)
    print("type counts:", dict(t_counts))
    print("match counts:", dict(m_counts))
    print("sample rows:")
    for r in records[:sample]:
        print(
            {
                "type": r.type,
                "match": r.match,
                "uuid": r.uuid,
                "naam": r.naam,
                "naampad": r.naampad,
                "isActief": r.isActief,
                "toestand": r.toestand,
                "datum": r.datum_laatste_keuring,
                "resultaat": r.resultaat_keuring,
            }
        )
    return 0


def run_smoke(settings_path: Path, env: Environment, ls_short_uri: str, lsdeel_short_uri: str, limit: int = 50) -> int:
    """Small end-to-end smoke test: generate excel to tmpfile and compare totals.

    Returns 0 on success, non-zero on mismatch or error.
    """
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "smoke_keuring.xlsx"
        _, recs = generate_excel(settings_path, env, ls_short_uri, lsdeel_short_uri, out_path=out, limit=limit)

        report = compare_records_vs_excel(recs, out, include_not_meegenomen=False)

        diffs = {k: v for k, v in report.items() if v["diff"] != 0}
        if diffs:
            print("SMOKE FAIL: discrepancies found:")
            for k, v in report.items():
                print(f"{k}: py={v['py']} excel={v['excel']} diff={v['diff']}")
            return 2

        print("SMOKE OK: python pivot totals match Excel pivot totals (Pivot sheet)")
    return 0


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Debug export helper")
    parser.add_argument("--mode", type=str, default="both", choices=["generate", "compare", "both", "pivot-debug", "quick-stats", "smoke"], help="Action to perform")
    default_settings = Path('/home/davidlinux/Documenten/AWV/resources/settings_SyncToArangoDB.json')
    parser.add_argument("--settings", type=Path, default=default_settings, help="Path to settings.json")
    parser.add_argument("--env", type=str, default="PRD", choices=[e.name for e in Environment])
    parser.add_argument("--ls-short-uri", type=str, default="lgc:installatie#LS")
    parser.add_argument("--lsdeel-short-uri", type=str, default="lgc:installatie#LSDeel")
    parser.add_argument("--out", type=Path, default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--sample", type=int, default=10)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    env = Environment[args.env]

    mode = args.mode
    settings = args.settings
    ls = args.ls_short_uri
    lsdeel = args.lsdeel_short_uri
    out = args.out
    limit = args.limit

    if mode == 'generate':
        path, recs = generate_excel(settings, env, ls, lsdeel, out_path=out, limit=limit)
        print(f"Wrote {len(recs)} rows to {path}")
        return 0

    if mode == 'compare':
        if out is None:
            raise SystemExit("--out is required for compare mode (point to an existing Excel file)")
        # need to fetch records to compare
        settings_obj = _load_settings(settings)
        db = _create_db_from_settings(settings_obj, env=env)
        recs = fetch_records(db, ls_short_uri=ls, lsdeel_short_uri=lsdeel, limit=limit)
        report = compare_records_vs_excel(recs, out, include_not_meegenomen=False)
        for k, v in report.items():
            print(f"{k}: py={v['py']} excel={v['excel']} diff={v['diff']}")
        return 0

    if mode == 'both':
        path, recs = generate_excel(settings, env, ls, lsdeel, out_path=out, limit=limit)
        print(f"Wrote {len(recs)} rows to {path}")
        report = compare_records_vs_excel(recs, path, include_not_meegenomen=False)
        mismatches = {k: v for k, v in report.items() if v['diff'] != 0}
        if mismatches:
            print("MISMATCHES:")
            for k, v in report.items():
                print(f"{k}: py={v['py']} excel={v['excel']} diff={v['diff']}")
            return 2
        print("OK: pivot totals match (Pivot sheet)")
        return 0

    if mode == 'pivot-debug':
        return run_pivot_debug(settings, env, ls, lsdeel, limit=limit)

    if mode == 'quick-stats':
        return run_quick_stats(settings, env, ls, lsdeel, limit=limit, sample=args.sample)

    if mode == 'smoke':
        return run_smoke(settings, env, ls, lsdeel, limit=limit or 50)

    raise SystemExit("Unknown mode")


if __name__ == '__main__':
    raise SystemExit(main())
