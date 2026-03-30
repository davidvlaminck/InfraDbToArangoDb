"""Export keuringsinfo (LS/Laagspanningsbord) from ArangoDB to an Excel workbook.

This script is intentionally placed in Analysis/ as requested.

Assumptions
- Assets have `ins.EMObject_*` fields as described in Analysis/spec.md.
- Assets have `toezichtgroep_key` (8 chars) referring to toezichtgroepen._key.

Behaviour
- The report focuses on a single asset type (configurable via --asset-short-uri).
- Adds a `type` column (e.g. 'Laagspanningsbord') to make the reported type explicit.

Excel output
- Separate sheets per toezichtgroep + "Andere".
- A "Niet meegenomen" sheet for assets that are active but have toestand
  verwijderd/overgedragen.
  - Assets with AIMDBStatus_isActief == false are NEVER considered (not exported,
    not in pivot, not in "Niet meegenomen").
- A "Pivot" sheet with counts per toezichtgroep (plus a total row) as rows and
  keuringsresultaat as columns. Keuringsresultaat only counts when
  datum_laatste_keuring > 2021-01-01; otherwise it's treated as blank.
"""

from __future__ import annotations

import argparse
import datetime as dt
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from API.APIEnums import AuthType, Environment
from ArangoDBConnectionFactory import ArangoDBConnectionFactory


TARGET_SHEETS = {"V&W-WL", "V&W-WA", "V&W-WO", "V&W-WW", "V&W-WVB", "Tunnel Organ. VL."}

# Map alternative toezichtgroep labels (case-insensitive) to the canonical sheet name
# e.g. some records use 'V&W Vlaams-Brabant' which should be treated as 'V&W-WVB'
SHEET_ALIASES: dict[str, str] = {
    "v&w vlaams-brabant": "V&W-WVB",
    "v&w-vlaams-brabant": "V&W-WVB",
    "v&w-wvb": "V&W-WVB",
    # Agent display-name variants (from agents collection) -> canonical sheet
    # confirmed mappings provided by the user
    "v&w oost-vlaanderen": "V&W-WO",
    "v&w limburg": "V&W-WL",
    "v&w antwerpen": "V&W-WA",
    "v&w west-vlaanderen": "V&W-WW",
    # Tunnel organisation (agent) maps to the Tunnel sheet
    "tunnel organ. vl.": "Tunnel Organ. VL.",
}

# Explicit mapping from agent document keys/uuids to canonical sheet names.
# The user provided these authoritative mappings (short _key -> display name -> uuid).
# We include both the short _key and the full uuid to maximise matching chance when
# assets store either form in their toezichtgroep/toezichtgroep_key fields.
AGENT_TO_SHEET: dict[str, str] = {
    # V&W agents
    "206ba12e-dcc6": "V&W-WO",  # V&W Oost-Vlaanderen (_key)
    "206ba12e-dcc6-4ed1-887c-978e98aaad41": "V&W-WO",  # uuid
    "4761b281-fe11": "V&W-WL",  # V&W Limburg
    "4761b281-fe11-4645-93a6-0e1955330e1c": "V&W-WL",
    "5efe6764-007f": "V&W-WA",  # V&W Antwerpen
    "5efe6764-007f-4099-83d9-29d0b2759211": "V&W-WA",
    "61f977f9-f8c6": "V&W-WW",  # V&W West-Vlaanderen
    "61f977f9-f8c6-4faf-859a-2cc180b61511": "V&W-WW",
    "e3fe5c8e-037b": "V&W-WVB",  # V&W Vlaams-Brabant
    "e3fe5c8e-037b-40cd-bff7-4617eb8bb86a": "V&W-WVB",

    # Tunnel organisation
    "7aa92dda-9e03": "Tunnel Organ. VL.",
    "7aa92dda-9e03-4f10-a0b3-1c6748c332b9": "Tunnel Organ. VL.",
}

EXCLUDED_SHEET = "Niet meegenomen"
PIVOT_SHEET = "Pivot"
PIVOT_ALL_SHEET = "Pivot (incl Niet meegenomen)"


@dataclass(frozen=True)
class KeuringsRecord:
    toezichtgroep: str
    type: str  # LS | Laagspanningsbord (LSB)
    # match field removed (always 'single' before)
    uuid: str
    betrokken_agent_key: str | None = None
    betrokken_agent_uuid: str | None = None
    betrokken_agent_org_key: str | None = None
    betrokken_agent_org_uuid: str | None = None
    betrokken_agent_org_name: str | None = None
    # role-based agents discovered via betrokkenerelaties
    toezichtgroep_agent_key: str | None = None
    toezichtgroep_agent_uuid: str | None = None
    toezichtgroep_agent_name: str | None = None
    toezichter_agent_key: str | None = None
    toezichter_agent_uuid: str | None = None
    toezichter_agent_name: str | None = None
    toezichtgroep_raw: str | None = None
    toezichtgroep_key_raw: str | None = None
    lsb_uuid: str | None = None
    naam: str | None = None
    naampad: str | None = None
    isActief: bool | None = None
    toestand: str | None = None
    datum_laatste_keuring: str | None = None
    resultaat_keuring: str | None = None
    longitude: float | None = None
    latitude: float | None = None


def _load_settings(path: Path) -> dict[str, Any]:
    import json

    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _create_db_from_settings(settings: dict[str, Any], env: Environment) -> Any:
    db_settings = settings["databases"][str(env.value)]
    factory = ArangoDBConnectionFactory(
        db_name=db_settings["database"],
        username=db_settings["user"],
        password=db_settings["password"],
    )
    return factory.create_connection()


def _parse_iso_date(s: str | None) -> dt.date | None:
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s)
    except ValueError:
        return None


def _load_technique_map(tree_structures_path: Path) -> dict[str, str]:
    """Load tree_structures.json and return a map lsdeel_uuid -> label.

    If the file doesn't exist or is malformed, return an empty map.
    """
    import json

    mapping: dict[str, str] = {}
    try:
        if not tree_structures_path.exists():
            return mapping
        data = json.loads(tree_structures_path.read_text(encoding="utf-8"))
    except Exception:
        return mapping

    # tree_structures.json can be a list or an object map
    entries = data if isinstance(data, list) else list(data.values())
    for e in entries:
        if not isinstance(e, dict):
            continue
        label = e.get("label") or ""
        # support both legacy key 'lsdeel_uuids' and new 'lsb_uuids'
        for u in (e.get("lsb_uuids") or []) + (e.get("lsdeel_uuids") or []):
            if isinstance(u, str) and u:
                mapping[u] = label
    return mapping


def build_aql(
    asset_short_uri: str,
    *,
    limit: int | None = None,
) -> str:
    """Build AQL to emit one row per asset of the given asset_short_uri.

    Simplified: we no longer consider LS + LSDeel pairing. The report only
    contains assets of the requested type (e.g. onderdeel#Laagspanningsbord).

    Inclusion: Only assets with AIMDBStatus_isActief == true are returned.
    Routing: Assets with toestand verwijderd/overgedragen are still returned,
    but will be placed on the "Niet meegenomen" sheet in Python.
    """

    # Always resolve the assettype key for the requested short_uri and filter by it
    let_clause = "LET at_key = FIRST(FOR at IN assettypes FILTER at.short_uri == @asset_short_uri LIMIT 1 RETURN at._key)\n"
    filter_clause = "  FILTER a.assettype_key == at_key\n"

    aql = (
        let_clause
        + "FOR a IN assets\n"
        + "  FILTER a.AIMDBStatus_isActief == true\n"
        + filter_clause
        + "  LET tz = FIRST(FOR t IN toezichtgroepen FILTER t._key == a.toezichtgroep_key LIMIT 1 RETURN t)\n"
        + "  // find latest ElektrischeKeuring via assetrelaties (relatietype HeeftKeuring)\n"
        + "  LET key_relatie_heeftkeuring = FIRST(FOR rel_type IN relatietypes FILTER rel_type.naam == 'HeeftKeuring' LIMIT 1 RETURN rel_type._key)\n"
        + "  LET latest_keuring = FIRST(\n"
        + "    FOR v, e IN 1..1 OUTBOUND a assetrelaties\n"
        + "      FILTER e.relatietype_key == key_relatie_heeftkeuring\n"
        + "      FILTER v.KeuringObject_keuringsdatum != null\n"
        + "      SORT v.KeuringObject_keuringsdatum DESC\n"
        + "      LIMIT 1\n"
        + "      RETURN v\n"
        + "  )\n"
        + "  LET __result_array = (latest_keuring != null && latest_keuring.ElektrischeKeuring_resultaat != null ? SPLIT(latest_keuring.ElektrischeKeuring_resultaat, '/') : null)\n"
        + "  LET resultaat_keuring_stripped = (__result_array != null ? __result_array[LENGTH(__result_array) - 1] : null)\n"
        + "  LET betrokken_relaties = (FOR v, e IN 1..1 OUTBOUND a betrokkenerelaties RETURN {edge: e, vertex: v})\n"
        + "  LET toezichtgroep_agent = FIRST(FOR rel IN betrokken_relaties FILTER rel.edge.rol == 'toezichtsgroep' RETURN rel.vertex)\n"
        + "  LET toezichter_agent = FIRST(FOR rel IN betrokken_relaties FILTER rel.edge.rol == 'toezichter' RETURN rel.vertex)\n"
        + "  LET toezichtgroep_agent_name = (toezichtgroep_agent != null && toezichtgroep_agent.purl != null && toezichtgroep_agent.purl.Agent_naam != null ? toezichtgroep_agent.purl.Agent_naam : (toezichtgroep_agent != null ? toezichtgroep_agent.naam : null))\n"
        + "  LET tz_from_agent = toezichtgroep_agent_name\n\n"
        + "  RETURN {\n"
        + "    \"toezichtgroep\": tz != null ? tz.naam : (tz_from_agent != null ? tz_from_agent : \"UNKNOWN\"),\n"
        + "    \"toezichtgroep_raw\": (a.toezichtgroep != null ? a.toezichtgroep : null),\n"
        + "    \"toezichtgroep_key_raw\": (a.toezichtgroep_key != null ? a.toezichtgroep_key : null),\n"
        + "    \"betrokken_agent_key\": null,\n"
        + "    \"betrokken_agent_uuid\": null,\n"
        + "    \"betrokken_agent_org_key\": null,\n"
        + "    \"betrokken_agent_org_uuid\": null,\n"
        + "    \"betrokken_agent_org_name\": null,\n"
        + "    \"toezichtgroep_agent_key\": (toezichtgroep_agent != null ? toezichtgroep_agent._key : null),\n"
        + "    \"toezichtgroep_agent_uuid\": (toezichtgroep_agent != null ? toezichtgroep_agent.uuid : null),\n"
        + "    \"toezichtgroep_agent_name\": (toezichtgroep_agent != null && toezichtgroep_agent.purl != null && toezichtgroep_agent.purl.Agent_naam != null ? toezichtgroep_agent.purl.Agent_naam : (toezichtgroep_agent != null ? toezichtgroep_agent.naam : null)),\n"
        + "    \"toezichter_agent_key\": (toezichter_agent != null ? toezichter_agent._key : null),\n"
        + "    \"toezichter_agent_uuid\": (toezichter_agent != null ? toezichter_agent.uuid : null),\n"
        + "    \"toezichter_agent_name\": (toezichter_agent != null && toezichter_agent.purl != null && toezichter_agent.purl.Agent_naam != null ? toezichter_agent.purl.Agent_naam : (toezichter_agent != null ? toezichter_agent.naam : null)),\n"
        + "    \"type\": @asset_short_uri == \"lgc:onderdeel#Laagspanningsbord\" ? \"Laagspanningsbord\" : @asset_short_uri,\n\n"
        + "    \"uuid\": a._key,\n"
        + "    \"lsb_uuid\": null,\n"
        + "    \"naam\": a.AIMNaamObject_naam,\n"
        + "    \"naampad\": a.NaampadObject_naampad,\n\n"
        + "    \"isActief\": a.AIMDBStatus_isActief,\n"
        + "    \"toestand\": a.toestand,\n\n"
        + "    \"datum_laatste_keuring\": (latest_keuring != null ? latest_keuring.KeuringObject_keuringsdatum : null),\n"
        + "    \"resultaat_keuring\": resultaat_keuring_stripped,\n\n"
        + "    \"longitude\": (a.geometry != null ? (LENGTH(a.geometry.coordinates) > 0 ? a.geometry.coordinates[0] : null) : null),\n"
        + "    \"latitude\": (a.geometry != null ? (LENGTH(a.geometry.coordinates) > 1 ? a.geometry.coordinates[1] : null) : null)\n"
        + "  }"
    )
    return aql


def fetch_records(
    db: Any,
    asset_short_uri: str,
    *,
    max_runtime_seconds: int = 300,
    limit: int | None = None,
) -> list[KeuringsRecord]:
    aql = build_aql(asset_short_uri=asset_short_uri)
    bind_vars: dict[str, Any] = {
        "asset_short_uri": asset_short_uri,
    }

    cursor = db.aql.execute(
        aql,
        bind_vars=bind_vars,
        batch_size=2000,
        ttl=600,
        max_runtime=max_runtime_seconds,
        stream=True,
    )

    if limit is None:
        return [KeuringsRecord(**row) for row in cursor]

    res = []
    for i, row in enumerate(cursor):
        if i >= limit:
            break
        res.append(KeuringsRecord(**row))
    return res


def fetch_records_not_meegenomen(*args: Any, **kwargs: Any) -> list[KeuringsRecord]:
    """Deprecated: kept for backward compatibility, but no longer used."""
    raise NotImplementedError("Not needed anymore: use fetch_records() and route by toestand in Python")


def _sheet_name(toezichtgroep: str | None) -> str:
    if not toezichtgroep:
        return "Andere"
    # normalize and apply known aliases (case-insensitive)
    key = toezichtgroep.strip()
    key_lower = key.lower()

    # If the toezichtgroep value directly references an agent _key or uuid,
    # the AGENT_TO_SHEET mapping takes precedence.
    if key in AGENT_TO_SHEET:
        return AGENT_TO_SHEET[key]
    if key_lower in AGENT_TO_SHEET:
        return AGENT_TO_SHEET[key_lower]

    # alias map first (case-insensitive)
    if key_lower in SHEET_ALIASES:
        return SHEET_ALIASES[key_lower]

    # case-insensitive match against canonical target names
    for ts in TARGET_SHEETS:
        if ts.lower() == key_lower:
            return ts

    return "Andere"


def _is_not_included(record: KeuringsRecord) -> bool:
    # In "Niet meegenomen":
    # - ONLY active assets with removed/transferred toestand
    # - Inactive assets are never exported at all
    return (record.toestand or "").lower() in {"verwijderd", "overgedragen"}


def _pivot_result_key(record: KeuringsRecord, *, cutoff: dt.date) -> str:
    """Return pivot category for this record.

    Robust rules (improved): normalize resultaat_keuring and match with substrings so
    variants such as 'niet-conform met inbreuken' or 'niet-conform (met inbreuken)'
    are handled correctly.

    Rules:
    - If no keuringsdatum: 'geen keuring'
    - If keuringsdatum <= cutoff:
        - If resultaat indicates conform (or conform met opmerkingen): 'vervallen keuring, conform'
        - If resultaat indicates niet-conform: 'vervallen keuring, niet conform'
        - Else: 'geen keuring'
    - If keuringsdatum > cutoff:
        - conform: 'conform'
        - conform met opmerkingen: 'conform met opmerkingen'
        - niet-conform: 'niet-conform met inbreuken'
        - Else: 'geen keuring'
    """
    d = _parse_iso_date(record.datum_laatste_keuring)
    r = record.resultaat_keuring
    r_norm = r.strip().lower() if r else None

    # helper normalizations
    def _is_not_conform(s: str) -> bool:
        return 'niet' in s and 'conform' in s or s.startswith('niet-')

    def _is_conform(s: str) -> bool:
        # ensure we don't misclassify 'niet-conform' as 'conform'
        return ('conform' in s) and (not _is_not_conform(s))

    def _has_opmerking(s: str) -> bool:
        return 'opmerking' in s or 'opmerk' in s

    if d is None:
        return 'geen keuring'

    if r_norm and ('niet gekend' in r_norm or r_norm == 'geen keuring'):
        r_norm = None

    if d <= cutoff:
        if r_norm:
            if _is_not_conform(r_norm):
                return 'vervallen keuring, niet conform'
            if _is_conform(r_norm):
                # conform or conform met opmerkingen
                if _has_opmerking(r_norm):
                    return 'vervallen keuring, conform'
                return 'vervallen keuring, conform'
        return 'geen keuring'

    # d > cutoff
    if r_norm:
        if _is_not_conform(r_norm):
            return 'niet-conform met inbreuken'
        if _has_opmerking(r_norm):
            return 'conform met opmerkingen'
        if _is_conform(r_norm):
            return 'conform'
    return 'geen keuring'


def _pivot_group_name(record: KeuringsRecord) -> str:
    """Map any record to one of the 6 target groups or 'Andere'.

    For pivoting we intentionally collapse everything outside TARGET_SHEETS into
    'Andere' so the pivot stays stable.
    """

    # Prefer the explicit toezichtgroep name; if missing or UNKNOWN, prefer
    # role-based agents discovered via betrokkenerelaties (toezichtgroep_agent_name,
    # then toezichter_agent_name). This mirrors the exporter mapping logic.
    resolved = _resolved_toezichtgroep(record)
    return _sheet_name(resolved)


def _resolved_toezichtgroep(record: KeuringsRecord) -> str | None:
    """Return the best-available toezichtgroep display name for this record.

    Priority:
    1. record.toezichtgroep if set and not 'UNKNOWN'
    2. record.toezichtgroep_agent_name
    3. record.toezichter_agent_name
    4. None
    """
    if record.toezichtgroep and str(record.toezichtgroep).strip().upper() != "UNKNOWN":
        return record.toezichtgroep
    if getattr(record, 'toezichtgroep_agent_name', None):
        return getattr(record, 'toezichtgroep_agent_name')
    return None


def _build_pivot(
    records: Iterable[KeuringsRecord],
    *,
    cutoff: dt.date,
    include_not_meegenomen: bool = False,
) -> tuple[list[str], dict[str, Counter[str]]]:
    """Build pivot data with fixed column order and new categories."""
    counters: dict[str, Counter[str]] = defaultdict(Counter)
    all_results: set[str] = set()

    for r in records:
        if _is_not_included(r) and (not include_not_meegenomen):
            continue
        res = _pivot_result_key(r, cutoff=cutoff)
        grp = _pivot_group_name(r)
        counters[grp][res] += 1
        all_results.add(res)

    # Fixed column order as requested
    result_cols = [
        'conform',
        'conform met opmerkingen',
        'niet-conform met inbreuken',
        'vervallen keuring, conform',
        'vervallen keuring, niet conform',
        'geen keuring',
    ]
    # Ensure all columns are present in all groups
    for c in counters.values():
        for col in result_cols:
            c.setdefault(col, 0)
    return result_cols, counters


def _write_pivot_sheet(
    wb: Any,
    *,
    sheet_name: str,
    records: list[KeuringsRecord],
    cutoff: dt.date,
    include_not_meegenomen: bool,
) -> None:
    """Create/overwrite a Pivot sheet."""

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    sh = wb.create_sheet(title=sheet_name, index=0)

    cols, counters = _build_pivot(
        records,
        cutoff=cutoff,
        include_not_meegenomen=include_not_meegenomen,
    )

    header: list[Any] = ["toezichtgroep", *cols, "Totaal"]
    sh.append(header)

    # fixed ordering: target groups (sorted) then Andere
    row_groups = [*sorted(TARGET_SHEETS), "Andere"]

    grand_total: Counter[str] = Counter()
    for tg in row_groups:
        c = counters.get(tg, Counter())
        row: list[Any] = [tg]
        row_total = 0
        for res in cols:
            v = int(c.get(res, 0))
            row.append(v)
            row_total += v
            grand_total[res] += v
        row.append(row_total)
        sh.append(row)

    total_row: list[Any] = ["Totaal"]
    total_sum = 0
    for res in cols:
        v = int(grand_total.get(res, 0))
        total_row.append(v)
        total_sum += v
    total_row.append(total_sum)
    sh.append(total_row)


def export_to_excel(records: Iterable[KeuringsRecord], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    import json

    # records may be a large iterable; avoid forcing it into a list here.
    # However other code relied on multiple passes, so we'll consume into a list
    # only if it's not already a list/tuple. Keep existing behavior but be defensive
    # about cell values (sanitize lists/dicts).
    records_list = list(records)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Two pivots:
    # - Pivot: excludes Niet meegenomen
    # - Pivot (incl Niet meegenomen): includes them
    cutoff = dt.date(2021, 1, 1)
    _write_pivot_sheet(
        wb,
        sheet_name=PIVOT_ALL_SHEET,
        records=records_list,
        cutoff=cutoff,
        include_not_meegenomen=True,
    )
    _write_pivot_sheet(
        wb,
        sheet_name=PIVOT_SHEET,
        records=records_list,
        cutoff=cutoff,
        include_not_meegenomen=False,
    )

    sheet_names = [*sorted(TARGET_SHEETS), "Andere", EXCLUDED_SHEET]
    sheets = {name: wb.create_sheet(title=name) for name in sheet_names}

    headers = [
        "toezichtgroep",
        "toezichter",
        "type",
        "uuid",
        "naam",
        "naampad",
        "techniek",
        "isActief",
        "toestand",
        "datum_laatste_keuring",
        "resultaat_keuring",
        "longitude",
        "latitude",
    ]

    for sh in sheets.values():
        sh.append(headers)

    # load technique mapping (lsdeel_uuid -> label) from tree_structures.json
    tree_structures_file = Path(__file__).parent / "TreeAnalysis" / "output" / "tree_structures.json"
    techniek_map = _load_technique_map(tree_structures_file)

    for r in records_list:
        if _is_not_included(r):
            sh = sheets[EXCLUDED_SHEET]
        else:
            # Resolve sheet name using several heuristics:
            # 1. Use the resolved toezichtgroep name (tz.naam) via _sheet_name
            # 2. If that yields 'Andere' or UNKNOWN, try to map using agent keys/uuids
            #    or raw toezichtgroep fields via AGENT_TO_SHEET and SHEET_ALIASES.
            resolved_name = _resolved_toezichtgroep(r)
            sheet_candidate = _sheet_name(resolved_name)
            if sheet_candidate == "Andere" or (r.toezichtgroep or "").upper() == "UNKNOWN":
                # try direct mapping by raw key/uuid
                mapped = None
                if getattr(r, 'toezichtgroep_key_raw', None):
                    k = r.toezichtgroep_key_raw
                    if k in AGENT_TO_SHEET:
                        mapped = AGENT_TO_SHEET[k]
                    elif k.lower() in AGENT_TO_SHEET:
                        mapped = AGENT_TO_SHEET[k.lower()]

                # try mapping by raw toezichtgroep string
                if mapped is None and getattr(r, 'toezichtgroep_raw', None):
                    raw = (r.toezichtgroep_raw or "").strip()
                    raw_lower = raw.lower()
                    if raw in AGENT_TO_SHEET:
                        mapped = AGENT_TO_SHEET[raw]
                    elif raw_lower in AGENT_TO_SHEET:
                        mapped = AGENT_TO_SHEET[raw_lower]
                    else:
                        # try substring match: agent key/uuid inside raw
                        for ak, sheetn in AGENT_TO_SHEET.items():
                            if ak in raw or ak in (r.toezichtgroep or ""):
                                mapped = sheetn
                                break

                # try mapping using role-based toezichtgroep_agent fields added by AQL
                if mapped is None:
                    for fld in ("toezichtgroep_agent_key", "toezichtgroep_agent_uuid", "toezichtgroep_agent_name"):
                        v = getattr(r, fld, None)
                        if not v:
                            continue
                        v_str = str(v)
                        if v_str in AGENT_TO_SHEET:
                            mapped = AGENT_TO_SHEET[v_str]
                            break
                        if v_str.lower() in AGENT_TO_SHEET:
                            mapped = AGENT_TO_SHEET[v_str.lower()]
                            break

                if mapped:
                    sh = sheets.get(mapped, sheets["Andere"])
                else:
                    sh = sheets[_sheet_name(resolved_name)]
            else:
                sh = sheets[sheet_candidate]

        # map techniek by lsb uuid (preferred). fallback to the chosen uuid.
        techniek = ""
        if getattr(r, 'lsb_uuid', None):
            techniek = techniek_map.get(r.lsb_uuid, "")
        if not techniek:
            techniek = techniek_map.get(r.uuid, "")

        # sanitize values before appending to avoid unsupported types (lists/dicts)
        def _sanitize(v):
            if v is None:
                return None
            # numeric types are fine
            if isinstance(v, (int, float, str, bool)):
                return v
            # coordinates sometimes arrive as list like [lon, lat, z]
            if isinstance(v, (list, tuple)):
                # if numeric coordinate list, return first numeric as longitude-like
                if all(isinstance(x, (int, float)) for x in v) and len(v) >= 1:
                    # join only if length > 2? prefer scalar for Excel
                    return v[0] if len(v) == 1 or len(v) >= 1 else ','.join(map(str, v))
                return json.dumps(v, ensure_ascii=False)
            if isinstance(v, dict):
                return json.dumps(v, ensure_ascii=False)
            # fallback to str
            return str(v)

        resolved_name = _resolved_toezichtgroep(r)
        row_values = [
            _sanitize(resolved_name if resolved_name is not None else r.toezichtgroep),
            _sanitize(getattr(r, 'toezichter_agent_name', None)),
            _sanitize(r.type),
            _sanitize(r.uuid),
            _sanitize(r.naam),
            _sanitize(r.naampad),
            _sanitize(techniek),
            _sanitize(r.isActief),
            _sanitize(r.toestand),
            _sanitize(r.datum_laatste_keuring),
            _sanitize(r.resultaat_keuring),
            _sanitize(r.longitude),
            _sanitize(r.latitude),
        ]
        sh.append(row_values)

    def _autofit_sheet_columns(ws: Any, *, min_width: int = 10, max_width: int = 80, padding: int = 2) -> None:
        """Auto-adjust column widths based on cell value length."""

        ws.freeze_panes = "A2"  # keep header row visible

        # compute widths
        for col_cells in ws.columns:
            # col_cells can include merged cells etc; guard for missing column index
            first = next(iter(col_cells), None)
            if first is None or getattr(first, "column", None) is None:
                continue

            max_len = 0
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)

            width = min(max(min_width, max_len + padding), max_width)
            col_letter = get_column_letter(first.column)
            ws.column_dimensions[col_letter].width = width

    # Auto-fit all sheets (including Pivot)
    for ws in wb.worksheets:
        _autofit_sheet_columns(ws)

    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export keuringsinfo naar Excel")
    parser.add_argument("--settings", type=Path, required=True)
    parser.add_argument("--env", type=str, default="PRD", choices=[e.name for e in Environment])
    parser.add_argument("--auth", type=str, default="JWT", choices=[a.name for a in AuthType])
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).with_name(f"keuringsinfo_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
    )

    parser.add_argument(
        "--asset-short-uri",
        type=str,
        default="onderdeel#Laagspanningsbord",
        help="assettypes.short_uri for the asset to report (e.g. onderdeel#Laagspanningsbord)",
    )

    parser.add_argument("--limit", type=int, default=None, help="Optional cap rows (debug)")

    args = parser.parse_args()

    settings = _load_settings(args.settings)
    db = _create_db_from_settings(settings, env=Environment[args.env])

    records = fetch_records(
        db,
        asset_short_uri=args.asset_short_uri,
        max_runtime_seconds=300,
        limit=args.limit,
    )

    export_to_excel(records, args.out)
    print(f"Wrote {len(records)} rows to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
