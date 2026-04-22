#!/usr/bin/env python3
"""Export one flat Excel sheet with active Laagspanningsborden and all active ElektrischeKeuringen.

One output row corresponds to one ElektrischeKeuring linked to a Laagspanningsbord via an
active `HeeftKeuring` relation. The Laagspanningsbord fields are repeated for every linked
keuring. If a Laagspanningsbord has no active keuring, it is still exported once with empty
keuring columns so the list remains complete.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from API.APIEnums import Environment
from Analysis.export_keuringsinfo import (
    _create_db_from_settings,
    _load_settings,
    _load_technique_map,
)

DEFAULT_SETTINGS_PATH = Path("/home/davidlinux/Documenten/AWV/resources/settings_SyncToArangoDB.json")
DEFAULT_LSB_SHORT_URI = "onderdeel#Laagspanningsbord"
DEFAULT_KEURING_SHORT_URI = "onderdeel#ElektrischeKeuring"
DEFAULT_RELATIETYPE_NAAM = "HeeftKeuring"
DEFAULT_SHEET_NAME = "Laagspanningsborden"

EXCLUDED_KEURING_COLUMNS = {
    "keuring_id",
    "keuring_rev",
    "keuring_type",
    "keuring_assettype_key",
    "keuring_AbstracteAanvullendeGeometrie_typeURI",
    "keuring_AbstracteAanvullendeGeometrie_assetId_DtcIdentificator_identificator",
    "keuring_AbstracteAanvullendeGeometrie_assetId_DtcIdentificator_toegekendDoor",
}

EXCLUDED_KEURING_PREFIXES = (
    "keuring_AIMVersie_assetVersie",
)


def build_aql() -> str:
    """Return the AQL for the detailed LSB + keuring export."""
    return (
        "LET lsb_key = FIRST(FOR at IN assettypes FILTER at.short_uri == @lsb_short_uri LIMIT 1 RETURN at._key)\n"
        "LET keuring_key = FIRST(FOR at IN assettypes FILTER at.short_uri == @keuring_short_uri LIMIT 1 RETURN at._key)\n"
        "LET heeft_keuring_key = FIRST(FOR rt IN relatietypes FILTER rt.naam == @relatietype_naam LIMIT 1 RETURN rt._key)\n"
        "FOR lsb IN assets\n"
        "  FILTER lsb.AIMDBStatus_isActief == true\n"
        "  FILTER lsb.assettype_key == lsb_key\n"
        "  LET betrokken_relaties = (\n"
        "    FOR v, e IN 1..1 OUTBOUND lsb betrokkenerelaties\n"
        "      FILTER e.AIMDBStatus_isActief == null || e.AIMDBStatus_isActief == true\n"
        "      RETURN {edge: e, vertex: v}\n"
        "  )\n"
        "  LET toezichtgroep_agent = FIRST(\n"
        "    FOR rel IN betrokken_relaties\n"
        "      FILTER rel.edge.rol == 'toezichtsgroep'\n"
        "      RETURN rel.vertex\n"
        "  )\n"
        "  LET toezichter_agent = FIRST(\n"
        "    FOR rel IN betrokken_relaties\n"
        "      FILTER rel.edge.rol == 'toezichter'\n"
        "      RETURN rel.vertex\n"
        "  )\n"
        "  LET toezichtgroep_name = (\n"
        "    toezichtgroep_agent != null && toezichtgroep_agent.purl != null && toezichtgroep_agent.purl.Agent_naam != null\n"
        "      ? toezichtgroep_agent.purl.Agent_naam\n"
        "      : (toezichtgroep_agent != null ? toezichtgroep_agent.naam : null)\n"
        "  )\n"
        "  LET toezichter_name = (\n"
        "    toezichter_agent != null && toezichter_agent.purl != null && toezichter_agent.purl.Agent_naam != null\n"
        "      ? toezichter_agent.purl.Agent_naam\n"
        "      : (toezichter_agent != null ? toezichter_agent.naam : null)\n"
        "  )\n"
        "  LET keuringen = (\n"
        "    FOR keuring, e IN 1..1 OUTBOUND lsb assetrelaties\n"
        "      FILTER e.relatietype_key == heeft_keuring_key\n"
        "      FILTER e.AIMDBStatus_isActief == null || e.AIMDBStatus_isActief == true\n"
        "      FILTER keuring.assettype_key == keuring_key\n"
        "      FILTER keuring.AIMDBStatus_isActief == true\n"
        "      SORT keuring.KeuringObject_keuringsdatum DESC, keuring._key ASC\n"
        "      RETURN keuring\n"
        "  )\n"
        "  FOR keuring IN (LENGTH(keuringen) > 0 ? keuringen : [null])\n"
        "    SORT lsb.NaampadObject_naampad ASC, keuring.KeuringObject_keuringsdatum DESC, keuring._key ASC\n"
        "    RETURN {\n"
        "      uuid: lsb._key,\n"
        "      naampad: lsb.NaampadObject_naampad,\n"
        "      toezichtgroep: toezichtgroep_name,\n"
        "      toezichter: toezichter_name,\n"
        "      keuring: keuring\n"
        "    }\n"
    )


def fetch_rows(
    db: Any,
    *,
    lsb_short_uri: str = DEFAULT_LSB_SHORT_URI,
    keuring_short_uri: str = DEFAULT_KEURING_SHORT_URI,
    relatietype_naam: str = DEFAULT_RELATIETYPE_NAAM,
    max_runtime_seconds: int = 600,
    limit: int | None = None,
) -> list[dict[str, Any]]:
    """Fetch flat base rows with nested keuring documents."""
    cursor = db.aql.execute(
        build_aql(),
        bind_vars={
            "lsb_short_uri": lsb_short_uri,
            "keuring_short_uri": keuring_short_uri,
            "relatietype_naam": relatietype_naam,
        },
        batch_size=2000,
        ttl=600,
        max_runtime=max_runtime_seconds,
        stream=True,
    )

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(cursor):
        if limit is not None and index >= limit:
            break
        rows.append(row)
    return rows


def _load_migration_map(migration_file: Path) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    if not migration_file.exists():
        return mapping

    data = json.loads(migration_file.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        return mapping

    for item in data:
        if not isinstance(item, dict):
            continue
        uuid = item.get("uuid")
        migrated = item.get("migrated_from_uuids") or []
        if uuid and isinstance(migrated, list):
            mapping[str(uuid)] = [str(v) for v in migrated if v]
    return mapping


def _resolve_techniek(
    lsb_uuid: str,
    *,
    migration_map: dict[str, list[str]],
    techniek_map: dict[str, str],
) -> str:
    migrated_from = migration_map.get(str(lsb_uuid), [])
    for legacy_uuid in migrated_from:
        label = techniek_map.get(legacy_uuid)
        if label:
            return label
    return techniek_map.get(str(lsb_uuid), "")


def _normalize_key_part(key: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-z_]+", "_", key).strip("_")
    return normalized or "value"


def _flatten_document(value: Any, *, prefix: str) -> dict[str, Any]:
    """Flatten nested dicts; keep lists as JSON strings to preserve full content."""
    flat: dict[str, Any] = {}

    def visit(current: Any, path: str) -> None:
        if isinstance(current, dict):
            if not current:
                flat[path] = "{}"
                return
            for child_key, child_value in current.items():
                next_path = f"{path}_{_normalize_key_part(str(child_key))}"
                visit(child_value, next_path)
            return

        if isinstance(current, list):
            flat[path] = json.dumps(current, ensure_ascii=False)
            return

        flat[path] = current

    visit(value, prefix)
    return flat


def _sanitize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _autofit_sheet_columns(ws: Any, *, min_width: int = 10, max_width: int = 80, padding: int = 2) -> None:
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        first = next(iter(col_cells), None)
        if first is None or getattr(first, "column", None) is None:
            continue

        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))

        width = min(max(min_width, max_len + padding), max_width)
        ws.column_dimensions[get_column_letter(first.column)].width = width


def _prepare_export_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    tree_structures_file = Path(__file__).parent / "TreeAnalysis" / "output" / "tree_structures.json"
    migration_file = Path(__file__).parent / "migration_LSDeel.json"
    techniek_map = _load_technique_map(tree_structures_file)
    migration_map = _load_migration_map(migration_file)

    fixed_headers = ["uuid", "naampad", "toezichtgroep", "toezichter", "techniek", "toestand"]
    dynamic_headers: set[str] = set()
    prepared_rows: list[dict[str, Any]] = []

    for row in rows:
        prepared = {
            "uuid": row.get("uuid"),
            "naampad": row.get("naampad"),
            "toezichtgroep": row.get("toezichtgroep"),
            "toezichter": row.get("toezichter"),
            "techniek": _resolve_techniek(
                str(row.get("uuid") or ""),
                migration_map=migration_map,
                techniek_map=techniek_map,
            ),
        }

        keuring = row.get("keuring")
        if keuring is not None:
            flattened_keuring = _flatten_document(keuring, prefix="keuring")
            if "keuring_AIMToestand_toestand" in flattened_keuring:
                prepared["toestand"] = flattened_keuring.pop("keuring_AIMToestand_toestand")

            # Drop unwanted keuring fields from the export.
            for key in list(flattened_keuring.keys()):
                if key in EXCLUDED_KEURING_COLUMNS or any(key.startswith(prefix) for prefix in EXCLUDED_KEURING_PREFIXES):
                    flattened_keuring.pop(key, None)

            prepared.update(flattened_keuring)
            dynamic_headers.update(flattened_keuring.keys())

        prepared_rows.append(prepared)

    headers = fixed_headers + sorted(dynamic_headers)
    return headers, prepared_rows


def export_rows_to_excel(rows: list[dict[str, Any]], out_path: Path, *, sheet_name: str = DEFAULT_SHEET_NAME) -> None:
    from openpyxl import Workbook

    headers, prepared_rows = _prepare_export_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)

    for row in prepared_rows:
        ws.append([_sanitize_cell(row.get(header)) for header in headers])

    _autofit_sheet_columns(ws)
    wb.save(out_path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export all active Laagspanningsborden with all active ElektrischeKeuringen")
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS_PATH)
    parser.add_argument("--env", type=str, default="PRD", choices=[e.name for e in Environment])
    parser.add_argument("--out", type=Path, default=Path(__file__).with_name(f"lsb_keuring_details_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"))
    parser.add_argument("--lsb-short-uri", type=str, default=DEFAULT_LSB_SHORT_URI)
    parser.add_argument("--keuring-short-uri", type=str, default=DEFAULT_KEURING_SHORT_URI)
    parser.add_argument("--relatietype-naam", type=str, default=DEFAULT_RELATIETYPE_NAAM)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--max-runtime-seconds", type=int, default=600)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    settings = _load_settings(args.settings)
    db = _create_db_from_settings(settings, env=Environment[args.env])

    rows = fetch_rows(
        db,
        lsb_short_uri=args.lsb_short_uri,
        keuring_short_uri=args.keuring_short_uri,
        relatietype_naam=args.relatietype_naam,
        max_runtime_seconds=args.max_runtime_seconds,
        limit=args.limit,
    )
    export_rows_to_excel(rows, args.out)

    unique_assets = len({row.get("uuid") for row in rows})
    print(f"Fetched {len(rows)} export rows for {unique_assets} Laagspanningsborden")
    print(f"Wrote Excel file to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

